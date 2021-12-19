import cv2
import sys
import numpy as np
import os.path
import easyocr
import datetime
import openpyxl


# INPUT
folderPath = '/home/xense/HDD/Workspaces/pythonWorkspace/ML/projectML/project/'
dataPath = folderPath + 'Data/'
modelPath = folderPath + 'modelFiles/'
allImages = os.listdir(dataPath)
# imgVidAdd = 0
logFile = 'log.xlsx'
header = ['Time', 'PlateNumber']
isImage = True

# Initialize the parameters
confThreshold = 0.2  # Confidence threshold
nmsThreshold = 0.2  # Non-maximum suppression threshold

reader = easyocr.Reader(['en'], model_storage_directory=modelPath)

inpWidth = 416  # Width of network's input image
inpHeight = 416  # Height of network's input image


# Load names of classes1
classesFile = modelPath + "classes.names"

classes = None
with open(classesFile, 'rt') as f:
    classes = f.read().rstrip('\n').split('\n')

# Give the configuration and weight files for the model and load the network using them.

modelConfiguration = modelPath + "darknet-yolov3.cfg"
modelWeights = modelPath + "plateDetection.weights"

net = cv2.dnn.readNetFromDarknet(modelConfiguration, modelWeights)
net.setPreferableBackend(cv2.dnn.DNN_BACKEND_OPENCV)
net.setPreferableTarget(cv2.dnn.DNN_TARGET_CPU)


def getTodaysDate():
    today = datetime.date.today()
    return str(today)


def getCurrentDateTime():
    time = datetime.datetime.now()
    time = time.strftime("%d/%m/%Y %H:%M:%S")
    return str(time)


# Get the names of the output layers


def getOutputsNames(net):
    # Get the names of all the layers in the network
    layersNames = net.getLayerNames()
    # Get the names of the output layers, i.e. the layers with unconnected outputs
    return [layersNames[i[0] - 1] for i in net.getUnconnectedOutLayers()]


# Log the entry in excel

def performLogging(text):
    plateText = ' '.join(text)
    worksht.append([getCurrentDateTime(), plateText])
    workbook.save(logFile)
    return

# Draw the predicted bounding box


def drawPred(classId, conf, left, top, right, bottom):
    onlyPlate = frame.copy()[top:bottom, left:right]
    plateOnly = cv2.cvtColor(onlyPlate, cv2.COLOR_BGR2GRAY)
    thr = cv2.adaptiveThreshold(
        plateOnly, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 59, 15)
    text = reader.readtext(thr, detail=0)
    performLogging(text)
    cv2.imwrite(outputPlate, thr)

    cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 3)
    # Display the label at the top of the bounding box
    cv2.imwrite(outputFile, frame.astype(np.uint8))


# Remove the bounding boxes with low confidence using non-maxima suppression
def postprocess(frame, outs):
    frameHeight = frame.shape[0]
    frameWidth = frame.shape[1]

    # Scan through all the bounding boxes output from the network and keep only the
    # ones with high confidence scores. Assign the box's class label as the class with the highest score.
    classIds = []
    confidences = []
    boxes = []
    for out in outs:
        for detection in out:
            # if detection[4]>0.001:
            scores = detection[5:]
            classId = np.argmax(scores)
            # if scores[classId]>confThreshold:
            confidence = scores[classId]
            if confidence > confThreshold:
                center_x = int(detection[0] * frameWidth)
                center_y = int(detection[1] * frameHeight)
                width = int(detection[2] * frameWidth)
                height = int(detection[3] * frameHeight)
                left = int(center_x - width / 2)
                top = int(center_y - height / 2)
                classIds.append(classId)
                confidences.append(float(confidence))
                boxes.append([left, top, width, height])

    # Perform non maximum suppression to eliminate redundant overlapping boxes with
    # lower confidences.
    indices = cv2.dnn.NMSBoxes(boxes, confidences, confThreshold, nmsThreshold)
    for i in indices:
        i = i[0]
        box = boxes[i]
        left = box[0]
        top = box[1]
        width = box[2]
        height = box[3]
        drawPred(classIds[i], confidences[i], left,
                 top, left + width, top + height)


##########################
# INIT WORKBOOK AND WORKSHEET
try:
    workbook = openpyxl.load_workbook(logFile)
    workbook.close()
except:
    workbook = openpyxl.Workbook()
    workbook.save(logFile)
    workbook.close()

workbook = openpyxl.load_workbook(logFile)
if('Sheet' in workbook.sheetnames):
    workbook.remove(workbook.worksheets[workbook.sheetnames.index('Sheet')])
if(not getTodaysDate() in workbook.sheetnames):
    worksht = workbook.create_sheet(getTodaysDate())
    worksht.append(header)
else:
    allSheets = workbook.worksheets
    allSheetNames = workbook.sheetnames
    worksht = allSheets[allSheetNames.index(getTodaysDate())]
##########################


for each in allImages:
    imgVidAdd = dataPath + each
    cap = cv2.VideoCapture(imgVidAdd)

    outputFile = ""
    outputPlate = ""
    videoWriter = None
    if(isImage):
        outputFile = imgVidAdd.split('.')[0] + '_output.png'
        outputPlate = imgVidAdd.split('.')[0] + 'pl.png'
    else:
        outputFile = imgVidAdd.split('.')[0] + '_output.avi'
        vid_writer = cv2.VideoWriter(outputFile, cv2.VideoWriter_fourcc('M', 'J', 'P', 'G'), 30, (round(
            cap.get(cv2.CAP_PROP_FRAME_WIDTH)), round(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))))

    while cv2.waitKey(1) < 0:

        # get frame from the video
        hasFrame, frame = cap.read()

        # Stop the program if reached end of video
        if not hasFrame:
            break

        # Create a 4D blob from a frame.
        blob = cv2.dnn.blobFromImage(
            frame, 1/255, (inpWidth, inpHeight), [0, 0, 0], swapRB=True, crop=False)

        # Sets the input to the network
        net.setInput(blob)

        # Runs the forward pass to get output of the output layers
        outs = net.forward(getOutputsNames(net))

        # Remove the bounding boxes with low confidence
        postprocess(frame, outs)

        cv2.imshow('output', frame)

    cv2.destroyAllWindows()
