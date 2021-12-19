import datetime
import openpyxl
import os

header = ['Time', 'PlateNumber']


def getTodaysDate():
    today = datetime.date.today()
    return str(today)


def getCurrentDateTime():
    time = datetime.datetime.now()
    time = time.strftime("%d/%m/%Y %H:%M:%S")
    return str(time)


##########################
# INIT WORKBOOK AND WORKSHEET
try:
    workbook = openpyxl.load_workbook('log.xlsx')
    workbook.close()
except:
    workbook = openpyxl.Workbook()
    workbook.save('log.xlsx')
    workbook.close()

workbook = openpyxl.load_workbook('log.xlsx')
if('Sheet' in workbook.sheetnames):
    workbook.remove(workbook.worksheets[workbook.sheetnames.index('Sheet')])
if(not getTodaysDate() in workbook.sheetnames):
    worksht = workbook.create_sheet(getTodaysDate())
    worksht.append(header)
else:
    allSheets = workbook.worksheets
    allSheetNames = workbook.sheetnames
    worksht = allSheets[allSheetNames.index(getTodaysDate())]
##########################

worksht.append([getCurrentDateTime(), 'Adil'])

workbook.save('log.xlsx')
workbook.close()
